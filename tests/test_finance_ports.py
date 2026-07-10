import tempfile
import unittest
from pathlib import Path

from sevent4.adapters.finance_filesystem import (
    FileBudgetExplorerInputRepository,
    FileFinanceFlowInputRepository,
    FileMoneyFlowInputRepository,
    HtmlFileWriter,
)
from sevent4.application.finance import publish_budget_explorer, publish_finance_flow, publish_money_flow
from sevent4.finance.build_budget_explorer import build_budget_explorer_from_files
from sevent4.finance.build_finance_flow import build_finance_flow_from_files, render_html as render_finance_flow_html
from sevent4.finance.build_money_flow import build_money_flow_from_files
from sevent4.domain.finance_flow import gba_links_from_summaries


class FinancePortsTest(unittest.TestCase):
    def test_budget_explorer_application_publishes_through_ports(self) -> None:
        class City:
            name = "Testville"
            source_dir = Path("data/cities/testville/source")

        class Repository:
            def load(self):
                return type(
                    "Input",
                    (),
                    {
                        "city": City(),
                        "headline": [{"year": "2024-25"}],
                        "civic_meta": {"caveats": []},
                        "civic_rows": [],
                        "budget_stages": [{"label": "Draft", "amount_cr": 17018.0}],
                    },
                )()

        class Writer:
            def __init__(self) -> None:
                self.html = ""

            def write_html(self, html: str) -> None:
                self.html = html

        writer = Writer()
        result = publish_budget_explorer(
            Repository(),
            writer,
            lambda city, headline, civic_meta, civic_rows, budget_stages=None, deflator_series=None: f"{city.name}:{headline[0]['year']}:{budget_stages[0]['amount_cr']}",
        )

        self.assertEqual(writer.html, "Testville:2024-25:17018.0")
        self.assertEqual(result.html, "Testville:2024-25:17018.0")

    def test_money_flow_application_publishes_through_ports(self) -> None:
        class City:
            name = "Testville"
            source_dir = Path("data/cities/testville/source")

        class Repository:
            def load(self):
                return type("Input", (), {"city": City()})()

        class Writer:
            def __init__(self) -> None:
                self.html = ""

            def write_html(self, html: str) -> None:
                self.html = html

        writer = Writer()
        result = publish_money_flow(Repository(), writer, lambda city: f"money:{city.name}")

        self.assertEqual(writer.html, "money:Testville")
        self.assertEqual(result.html, "money:Testville")

    def test_finance_flow_application_publishes_through_ports(self) -> None:
        class City:
            name = "Testville"
            source_dir = Path("data/cities/testville/source")

        class Repository:
            def load(self):
                return type(
                    "Input",
                    (),
                    {
                        "city": City(),
                        "title": "Test flow",
                        "subtitle": "Fixture subtitle",
                        "links": [{"source": "A", "target": "B", "amount_cr": 12.5}],
                        "notes": ["source note"],
                    },
                )()

        class Writer:
            def __init__(self) -> None:
                self.html = ""

            def write_html(self, html: str) -> None:
                self.html = html

        writer = Writer()
        result = publish_finance_flow(
            Repository(),
            writer,
            lambda city, title, subtitle, links, notes, flow_years=None, default_year=None: f"{city.name}:{title}:{links[0]['amount_cr']}",
        )

        self.assertEqual(writer.html, "Testville:Test flow:12.5")
        self.assertEqual(result.html, "Testville:Test flow:12.5")

    def test_budget_explorer_file_repository_loads_city_and_budget_inputs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            city_yaml, _source_dir, _out = _write_finance_fixture(Path(tmp))

            inputs = FileBudgetExplorerInputRepository(city_yaml).load()

            self.assertEqual(inputs.city.name, "Testville")
            self.assertEqual(inputs.headline[0]["year"], "2023-24")
            self.assertEqual(inputs.civic_meta["caveats"][0], "fixture caveat")
            self.assertEqual(inputs.civic_rows[0]["line"], "AMTS")

    def test_finance_clis_build_html_through_file_adapters(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            city_yaml, _source_dir, out_dir = _write_finance_fixture(Path(tmp))
            budget_out = out_dir / "finance" / "index.html"
            money_out = out_dir / "money" / "index.html"
            flow_out = out_dir / "finance-flow" / "index.html"

            build_budget_explorer_from_files(str(city_yaml), str(budget_out))
            build_money_flow_from_files(str(city_yaml), str(money_out))
            build_finance_flow_from_files(str(city_yaml), str(flow_out), _source_dir / "gba_raw")

            budget_html = budget_out.read_text(encoding="utf-8")
            money_html = money_out.read_text(encoding="utf-8")
            flow_html = flow_out.read_text(encoding="utf-8")
            self.assertIn("What Testville city budget funds", budget_html)
            self.assertIn("Who controls the money", money_html)
            self.assertIn("GBA 2026-27 corporation budgets", flow_html)
            self.assertIn("Central", flow_html)
            self.assertIn("Receipts", flow_html)
            for html in (budget_html, money_html):
                self.assertNotIn('data-theme="dark"', html)
                self.assertIn('href="../../../assets/theme.css"', html)
                self.assertIn("localStorage.getItem('atlas-theme')", html)
            self.assertIn("@media (prefers-color-scheme: light)", money_html)
            self.assertIn(":root:not([data-theme=dark])", money_html)
            self.assertIn(":root[data-theme=light]", money_html)
            self.assertIn('href="../../../assets/theme.css"', flow_html)
            self.assertIn('href="../../../assets/masthead.css"', flow_html)
            self.assertIn('src="../../../assets/theme.js"', flow_html)
            self.assertIn('data-masthead="bar"', flow_html)

    def test_html_file_writer_creates_parent_directories(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "nested" / "index.html"

            HtmlFileWriter(out).write_html("hello")

            self.assertEqual(out.read_text(encoding="utf-8"), "hello")

    def test_money_flow_file_repository_loads_city(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            city_yaml, _source_dir, _out = _write_finance_fixture(Path(tmp))

            inputs = FileMoneyFlowInputRepository(city_yaml).load()

            self.assertEqual(inputs.city.name, "Testville")

    def test_finance_flow_file_repository_loads_gba_workbook_links(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            city_yaml, source_dir, _out = _write_finance_fixture(Path(tmp))

            inputs = FileFinanceFlowInputRepository(city_yaml, source_dir / "gba_raw").load()

            self.assertEqual(inputs.city.name, "Testville")
            self.assertIn("GBA 2026-27", inputs.title)
            labels = {(link["source"], link["target"]) for link in inputs.links}
            self.assertIn(("Central", "Receipts"), labels)
            self.assertIn(("Central", "Payments"), labels)
            self.assertTrue(any(link["amount_cr"] == 100.0 for link in inputs.links))

    def test_budget_explorer_file_repository_fills_ahmedabad_headline_gaps_from_budget_lines(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            city_yaml = _write_ahmedabad_budget_line_fixture(Path(tmp))

            inputs = FileBudgetExplorerInputRepository(city_yaml).load()

            by_year = {row["year"]: row for row in inputs.headline}
            self.assertEqual(by_year["2020-21"]["amts_cr"], 355.0)
            self.assertEqual(by_year["2020-21"]["mj_library_cr"], 16.691)
            self.assertEqual(by_year["2020-21"]["confidence"], "high")
            self.assertEqual(by_year["2022-23"]["amts_cr"], 390.0)
            self.assertEqual(by_year["2023-24"]["mj_library_cr"], 15.209)
            self.assertEqual(by_year["2026-27"]["total_cr"], 10500.09)
            self.assertEqual(inputs.budget_stages[0]["amount_cr"], 17018.0)
            self.assertEqual(inputs.budget_stages[1]["amount_cr"], 18518.0)

    def test_budget_explorer_renderer_keeps_total_budget_stage_separate_from_revenue_series(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            city_yaml = _write_ahmedabad_budget_line_fixture(Path(tmp))
            out = Path(tmp) / "public" / "cities" / "ahmedabad" / "finance" / "index.html"

            build_budget_explorer_from_files(str(city_yaml), str(out))

            html = out.read_text(encoding="utf-8")
            self.assertIn("2026-27 budget stages", html)
            self.assertIn("Draft total budget", html)
            self.assertIn("&#8377;17,018 cr", html)
            self.assertIn("Standing committee budget", html)
            self.assertIn("&#8377;18,518 cr", html)
            self.assertIn("<th>Total revenue budget</th>", html)
            self.assertNotIn("<th>Total budget</th>", html)

    def test_finance_flow_file_repository_loads_ahmedabad_budget_line_links(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            city_yaml = _write_ahmedabad_budget_line_fixture(Path(tmp))

            inputs = FileFinanceFlowInputRepository(city_yaml).load()

            self.assertEqual(inputs.city.name, "Ahmedabad")
            self.assertIn("Ahmedabad budget-line flow", inputs.title)
            labels = {(link["source"], link["target"]) for link in inputs.links}
            self.assertIn(("Ahmedabad Municipal Corporation", "Capital expenditure"), labels)
            self.assertIn(("Capital expenditure", "Loan/support to AMTS (city bus)"), labels)
            self.assertIn(("Revenue expenditure", "Grant to Municipal School Board"), labels)
            self.assertTrue(any("budget_line.csv" in note for note in inputs.notes))
            self.assertEqual(inputs.default_year, "2023-24")
            self.assertEqual([flow["year"] for flow in inputs.flow_years], ["2020-21", "2022-23", "2023-24"])
            self.assertTrue(any(flow["status"] == "partial" for flow in inputs.flow_years))
            self.assertTrue(all(flow["links"] for flow in inputs.flow_years))

    def test_finance_flow_renderer_emits_year_switcher_when_year_flows_exist(self) -> None:
        class City:
            name = "Ahmedabad"

        html = render_finance_flow_html(
            City(),
            "Ahmedabad budget-line flow",
            "Fixture subtitle",
            [{"source": "Corporation", "target": "Capital expenditure", "amount_cr": 10.0, "group": "payments"}],
            ["Fixture note"],
            flow_years=[
                {
                    "year": "2020-21",
                    "status": "partial",
                    "links": [{"source": "Corporation", "target": "Capital expenditure", "amount_cr": 8.0, "group": "payments"}],
                    "notes": ["2020 note"],
                    "rows": [],
                },
                {
                    "year": "2023-24",
                    "status": "complete",
                    "links": [{"source": "Corporation", "target": "Capital expenditure", "amount_cr": 10.0, "group": "payments"}],
                    "notes": ["2023 note"],
                    "rows": [],
                },
            ],
            default_year="2023-24",
        )

        self.assertIn('class="year-switcher"', html)
        self.assertIn('data-year="2020-21"', html)
        self.assertIn('data-year="2023-24"', html)
        self.assertIn('"defaultYear": "2023-24"', html)
        self.assertIn('"status": "partial"', html)
        self.assertIn("renderFinanceFlow", html)

    def test_gba_links_aggregate_payment_heads_for_readable_sankey(self) -> None:
        links = gba_links_from_summaries(
            [
                {
                    "corporation": "Central",
                    "total_receipts_cr": 100,
                    "total_payments_cr": 90,
                    "top_payment_heads": [
                        {"label": "Public Works", "amount_cr": 40},
                        {"label": "Solid Waste Management", "amount_cr": 20},
                    ],
                },
                {
                    "corporation": "East",
                    "total_receipts_cr": 80,
                    "total_payments_cr": 70,
                    "top_payment_heads": [
                        {"label": "Public Works", "amount_cr": 30},
                        {"label": "Health", "amount_cr": 10},
                    ],
                },
            ]
        )

        head_links = [link for link in links if link["source"] == "Payments"]

        self.assertEqual(head_links[0]["target"], "Public Works")
        self.assertEqual(head_links[0]["amount_cr"], 70)
        self.assertEqual(len([link for link in head_links if link["target"] == "Public Works"]), 1)

    def test_finance_flow_renderer_outputs_accessible_svg_and_table(self) -> None:
        class City:
            name = "Testville"

        html = render_finance_flow_html(
            City(),
            "Fixture Flow",
            "Fixture subtitle",
            [
                {"source": "A", "target": "B", "amount_cr": 10.0, "group": "receipts"},
                {"source": "A", "target": "C", "amount_cr": 6.0, "group": "payments"},
            ],
            ["Fixture note"],
        )

        self.assertIn('class="sankey"', html)
        self.assertIn('class="viz-scroll"', html)
        self.assertIn('role="img"', html)
        self.assertIn("band-receipts", html)
        self.assertIn("--flow-receipts", html)
        self.assertIn("fill:var(--flow-receipts)", html)
        self.assertNotIn('<path class="band" d=', html)
        self.assertIn("Fixture Flow", html)
        self.assertIn("Fixture note", html)
        self.assertIn("<table", html)

    def test_finance_flow_renderer_places_intermediate_nodes_in_middle_column(self) -> None:
        class City:
            name = "Testville"

        html = render_finance_flow_html(
            City(),
            "Fixture Flow",
            "Fixture subtitle",
            [
                {"source": "Corporation", "target": "Capital expenditure", "amount_cr": 10.0, "group": "payments"},
                {"source": "Capital expenditure", "target": "Roads", "amount_cr": 10.0, "group": "payment_head"},
            ],
            [],
        )

        self.assertIn('x="420.0" y="56.0" width="210.0"', html)
        self.assertIn(">Capital expenditure</text>", html)


def _write_finance_fixture(base: Path) -> tuple[Path, Path, Path]:
    repo = base / "repo"
    city_dir = repo / "data" / "cities" / "testville"
    source_dir = city_dir / "source"
    budget_dir = source_dir / "budget"
    out_dir = repo / "public" / "cities" / "testville"
    budget_dir.mkdir(parents=True)
    out_dir.mkdir(parents=True)
    (repo / "pyproject.toml").write_text("[project]\nname='fake'\n", encoding="utf-8")
    (city_dir / "city.yaml").write_text(
        "\n".join(
            [
                "id: testville",
                "name: Testville",
                "country: India",
                "state: State",
                "center: [72.0, 23.0]",
                "bbox: [71.0, 22.0, 73.0, 24.0]",
                "crs_metric: EPSG:32643",
                "layers_dir: data/cities/testville/layers",
                "source_dir: data/cities/testville/source",
                "outputs_dir: public/cities/testville",
            ]
        ),
        encoding="utf-8",
    )
    (budget_dir / "amc_budget_22yr.csv").write_text(
        "\n".join(
            [
                "year,amts_cr,mj_library_cr,property_tax_cr,total_cr,confidence,amts_page,notes",
                "2023-24,100,1,50,1000,high,10,",
                "2024-25,120,1.2,60,1100,high,11,",
            ]
        ),
        encoding="utf-8",
    )
    (budget_dir / "amc_civic_lines.json").write_text(
        '{"_meta":{"caveats":["fixture caveat"]},"data":[{"year":"2024-25","line":"AMTS","amount_cr":120}]}',
        encoding="utf-8",
    )
    _write_gba_workbook(source_dir / "gba_raw" / "Bengaluru_Central_City_Corporation_Budget_Tables.xlsx")
    return city_dir / "city.yaml", source_dir, out_dir


def _write_gba_workbook(path: Path) -> None:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "1_Financial_Position"
    ws.append(["BENGALURU CENTRAL CITY CORPORATION"])
    ws.append(["Financial Position 2026-27  (Rs. in Lakhs)"])
    ws.append(["Ref", "Particulars", "Budget Estimate\n2025-26", "Budget Estimate\n2026-27"])
    ws.append(["1 B1", "Revenue Receipts", 90, 6000])
    ws.append(["2 B2", "Revenue Payments", 70, 4000])
    ws.append(["1 C1", "Capital Receipts", 30, 4000])
    ws.append(["2 C2", "Capital Payments", 50, 5000])
    ws2 = wb.create_sheet("3_Receipts_Abstract")
    ws2.append(["BENGALURU CENTRAL CITY CORPORATION"])
    ws2.append(["Department-wise Abstract - RECEIPTS  (Rs. in Lakhs)"])
    ws2.append(["Sl. No.", "Department", "Budget Estimate\n2025-26", "Budget Estimate\n2026-27"])
    ws2.append(["1", "Revenue", 40, 7000])
    ws2.append(["2", "Town Planning", 20, 3000])
    ws2.append([None, "Total", 60, 10000])
    ws3 = wb.create_sheet("4_Payments_Abstract")
    ws3.append(["BENGALURU CENTRAL CITY CORPORATION"])
    ws3.append(["Function-wise Abstract - PAYMENTS  (Rs. in Lakhs)"])
    ws3.append(["Sl. No.", "Department", "Budget Estimate\n2025-26", "Budget Estimate\n2026-27"])
    ws3.append(["1", "Public Works", 45, 5500])
    ws3.append(["2", "Solid Waste Management", 20, 4500])
    ws3.append([None, "Total", 65, 10000])
    wb.save(path)


def _write_ahmedabad_budget_line_fixture(base: Path) -> Path:
    repo = base / "repo"
    city_dir = repo / "data" / "cities" / "ahmedabad"
    source_dir = city_dir / "source"
    budget_dir = source_dir / "budget"
    db_dir = city_dir / "db"
    budget_dir.mkdir(parents=True)
    db_dir.mkdir(parents=True)
    (repo / "pyproject.toml").write_text("[project]\nname='fake'\n", encoding="utf-8")
    (city_dir / "city.yaml").write_text(
        "\n".join(
            [
                "id: ahmedabad",
                "name: Ahmedabad",
                "country: India",
                "state: Gujarat",
                "center: [72.58, 23.03]",
                "bbox: [72.45, 22.90, 72.74, 23.18]",
                "crs_metric: EPSG:32643",
                "layers_dir: data/cities/ahmedabad/layers",
                "source_dir: data/cities/ahmedabad/source",
                "outputs_dir: public/cities/ahmedabad",
            ]
        ),
        encoding="utf-8",
    )
    (budget_dir / "amc_budget_22yr.csv").write_text(
        "\n".join(
            [
                "year,amts_cr,mj_library_cr,property_tax_cr,total_cr,confidence,amts_page,notes",
                "2019-20,339,12.8635,580.61,5196.93,high,138,",
                "2026-27,525,12,,,medium,299,",
            ]
        ),
        encoding="utf-8",
    )
    (budget_dir / "amc_civic_lines.json").write_text(
        '{"_meta":{"caveats":["fixture caveat"]},"data":[]}',
        encoding="utf-8",
    )
    (db_dir / "budget_line.csv").write_text(
        "\n".join(
            [
                "id,city,fiscal_year,fy_start,estimate_basis,section,flow,head_category,head_name,head_name_raw,entity,amount_cr,amount_raw,source_pdf,page,extraction_method,confidence,note",
                "1,ahmedabad,2020-21,2020,BE,capital,expenditure,department_support,Loan/support to AMTS (city bus),raw,AMTS,355,raw,AMC_Budget_2020-21_English7550.pdf,130,manual_verified,high,note",
                "2,ahmedabad,2020-21,2020,BE,revenue,expenditure,grant_contribution,Grant to Sheth M.J. Library,raw,MJ_LIBRARY,16.691,raw,AMC_Budget_2020-21_English7550.pdf,26,manual_verified,high,note",
                "3,ahmedabad,2022-23,2022,BE,capital,expenditure,department_support,Loan/support to AMTS (city bus),raw,AMTS,390,raw,AMC_Budget_2022-23_English2679.pdf,131,manual_verified,high,note",
                "8,ahmedabad,2022-23,2022,BE,revenue,expenditure,grant_contribution,Grant to Municipal School Board,raw,SCHOOL_BOARD,362,raw,AMC_Budget_2022-23_English2679.pdf,18,manual_verified,high,note",
                "9,ahmedabad,2022-23,2022,BE,capital,expenditure,capital_works,Sabarmati Riverfront (SRFDCL),raw,SRFDCL,84,raw,AMC_Budget_2022-23_English2679.pdf,148,manual_verified,high,note",
                "10,ahmedabad,2026-27,2026,BE,revenue,income,total,Total revenue budget,raw,,10500.09,raw,AMC_Budget_2026-27_Gujarati.pdf,33,manual_verified,high,note",
                "4,ahmedabad,2023-24,2023,BE,capital,expenditure,department_support,Loan/support to AMTS (city bus),raw,AMTS,398,raw,AMC_Budget_2023-24_English (1)6444.pdf,174,manual_verified,high,note",
                "5,ahmedabad,2023-24,2023,BE,revenue,expenditure,grant_contribution,Grant to Sheth M.J. Library,raw,MJ_LIBRARY,15.209,raw,AMC_Budget_2023-24_English (1)6444.pdf,28,manual_verified,high,note",
                "6,ahmedabad,2023-24,2023,BE,revenue,expenditure,grant_contribution,Grant to Municipal School Board,raw,SCHOOL_BOARD,506,raw,AMC_Budget_2023-24_English (1)6444.pdf,28,manual_verified,high,note",
                "7,ahmedabad,2023-24,2023,BE,capital,expenditure,capital_works,Sabarmati Riverfront (SRFDCL),raw,SRFDCL,160,raw,AMC_Budget_2023-24_English (1)6444.pdf,188,manual_verified,high,note",
            ]
        ),
        encoding="utf-8",
    )
    (budget_dir / "amc_budget_stages.csv").write_text(
        "\n".join(
            [
                "year,label,amount_cr,stage,source_label,source_url,note,confidence",
                "2026-27,Draft total budget,17018,draft,Times of India,https://timesofindia.indiatimes.com/city/ahmedabad/within-24-hours-of-presenting-proposal-in-2026-27-draft-budget-ahmedabad-municipal-corporation-axes-rs-32-crore-plan-to-relocate-mirzapur-slaughterhouse/articleshow/127950418.cms,reported commissioner draft budget,medium",
                "2026-27,Standing committee budget,18518,standing_committee,Times of India,https://timesofindia.indiatimes.com/city/ahmedabad/amc-sets-big-city-goals-with-rs-18518cr-blueprint-promises-beggar-free-ahmedabad/articleshow/128172537.cms,reported standing committee approved budget,medium",
            ]
        ),
        encoding="utf-8",
    )
    return city_dir / "city.yaml"


if __name__ == "__main__":
    unittest.main()
