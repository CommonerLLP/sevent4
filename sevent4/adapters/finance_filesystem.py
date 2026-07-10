from __future__ import annotations

import json
import re
from pathlib import Path

from sevent4.city_dataset import CityDataset
from sevent4.domain.finance_flow import amc_flow_year, amc_links_from_budget_lines, amc_year_flows, gba_flow_notes, gba_links_from_summaries
from sevent4.finance.budget_data import (
    civic_rows_from_budget_lines,
    enrich_headline_from_budget_lines,
    load_budget_lines,
    load_budget_stages,
    load_civic_lines,
    load_headline,
)
from sevent4.ports.finance import BudgetExplorerInput, FinanceFlowInput, MoneyFlowInput

ROOT = Path(__file__).resolve().parents[2]
GBA_RAW = ROOT / "data" / "sources" / "opencity" / "bengaluru" / "raw" / "gba-corporation-budgets-2026-27"
DEFLATOR_SERIES_PATH = ROOT / "data" / "references" / "deflator" / "cpi_combined_fy2005_06_to_latest.json"


def load_deflator_series(path: Path = DEFLATOR_SERIES_PATH) -> dict[str, float]:
    """Read the CPI series vendored from public-finance (REQ-0011). Pure IO,
    no computation — the deflate() logic lives in sevent4.domain.deflator."""
    payload = json.loads(path.read_text(encoding="utf-8"))
    return {row["fy"]: float(row["index_value"]) for row in payload["years"]}


class FileBudgetExplorerInputRepository:
    def __init__(self, city_config: str | Path) -> None:
        self.city_config = Path(city_config)

    def load(self) -> BudgetExplorerInput:
        city = CityDataset.from_yaml(self.city_config)
        budget_dir = city.source_dir / "budget"
        civic_meta, civic_rows = load_civic_lines(budget_dir / "amc_civic_lines.json")
        headline = load_headline(budget_dir / "amc_budget_22yr.csv")
        budget_stages = load_budget_stages(budget_dir / "amc_budget_stages.csv")
        budget_line_path = city.repo_root / "data" / "cities" / city.id / "db" / "budget_line.csv"
        if city.id == "ahmedabad" and budget_line_path.exists():
            budget_lines = load_budget_lines(budget_line_path)
            headline = enrich_headline_from_budget_lines(headline, budget_lines)
            civic_rows = civic_rows_from_budget_lines(budget_lines)
        return BudgetExplorerInput(
            city=city,
            headline=headline,
            civic_meta=civic_meta,
            civic_rows=civic_rows,
            budget_stages=budget_stages,
            deflator_series=load_deflator_series(),
        )


class FileMoneyFlowInputRepository:
    def __init__(self, city_config: str | Path) -> None:
        self.city_config = Path(city_config)

    def load(self) -> MoneyFlowInput:
        return MoneyFlowInput(city=CityDataset.from_yaml(self.city_config))


class FileFinanceFlowInputRepository:
    def __init__(self, city_config: str | Path, raw_dir: str | Path | None = None) -> None:
        self.city_config = Path(city_config)
        self.raw_dir = Path(raw_dir) if raw_dir is not None else GBA_RAW

    def load(self) -> FinanceFlowInput:
        city = CityDataset.from_yaml(self.city_config)
        budget_line_path = city.repo_root / "data" / "cities" / city.id / "db" / "budget_line.csv"
        if city.id == "ahmedabad" and budget_line_path.exists():
            budget_lines = load_budget_lines(budget_line_path)
            flow_years = amc_year_flows(budget_lines)
            year = amc_flow_year(budget_lines)
            current = next((flow for flow in flow_years if flow["year"] == year), {"links": [], "notes": []})
            return FinanceFlowInput(
                city=city,
                title="Ahmedabad budget-line flow over time",
                subtitle="Selected source-backed AMC expenditure lines from the canonical budget database",
                links=current["links"] or amc_links_from_budget_lines(budget_lines, year),
                notes=current["notes"],
                flow_years=flow_years,
                default_year=year,
            )
        summaries = read_gba_budget_summaries(self.raw_dir)
        links = gba_links_from_summaries(summaries)
        return FinanceFlowInput(
            city=city,
            title="GBA 2026-27 corporation budgets",
            subtitle="Five new Bengaluru corporations, from OpenCity budget-table workbooks",
            links=links,
            notes=gba_flow_notes(summaries),
        )


class HtmlFileWriter:
    def __init__(self, path: str | Path) -> None:
        self.path = Path(path)

    def write_html(self, html: str) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.path.write_text(html, encoding="utf-8")


def read_gba_budget_summaries(raw_dir: Path) -> list[dict]:
    from openpyxl import load_workbook

    summaries = []
    for workbook in sorted(raw_dir.glob("*.xlsx")):
        wb = load_workbook(workbook, read_only=True, data_only=True)
        corporation = _corporation_label(workbook.stem, wb[wb.sheetnames[0]].cell(1, 1).value)
        financial = _sheet_rows(wb["1_Financial_Position"])
        summaries.append(
            {
                "corporation": corporation,
                "source_file": workbook.name,
                "revenue_receipts_cr": _particular_amount(financial, "revenue receipts"),
                "capital_receipts_cr": _particular_amount(financial, "capital receipts"),
                "revenue_payments_cr": _particular_amount(financial, "revenue payments"),
                "capital_payments_cr": _particular_amount(financial, "capital payments"),
                "total_receipts_cr": _particular_amount(financial, "revenue receipts") + _particular_amount(financial, "capital receipts"),
                "total_payments_cr": _particular_amount(financial, "revenue payments") + _particular_amount(financial, "capital payments"),
                "top_payment_heads": _top_abstract_heads(wb, "payments"),
            }
        )
    return summaries


def _sheet_rows(sheet) -> list[tuple]:
    return [tuple(cell for cell in row) for row in sheet.iter_rows(values_only=True)]


def _corporation_label(stem: str, title_cell) -> str:
    text = " ".join(str(part) for part in (title_cell, stem) if part)
    for label in ("Central", "South", "East", "West", "North"):
        if label.upper() in text.upper():
            return label
    return stem.replace("_", " ")


def _budget_column(row: tuple) -> int:
    for index, value in enumerate(row):
        if isinstance(value, str) and "budget estimate" in value.lower() and "2026-27" in value:
            return index
    return len(row) - 1


def _particular_amount(rows: list[tuple], needle: str) -> float:
    amount_col = None
    for row in rows:
        if amount_col is None and any(isinstance(value, str) and "2026-27" in value for value in row):
            amount_col = _budget_column(row)
        if len(row) < 2 or row[1] is None:
            continue
        if needle in str(row[1]).lower():
            value = row[amount_col if amount_col is not None else len(row) - 1]
            return round(float(value or 0) / 100, 4)
    return 0.0


def _top_abstract_heads(workbook, kind: str) -> list[dict]:
    sheet = next((workbook[name] for name in workbook.sheetnames if kind.lower() in name.lower() and "abstract" in name.lower()), None)
    if sheet is None:
        return []
    rows = _sheet_rows(sheet)
    header_index = next(
        index for index, row in enumerate(rows)
        if any(isinstance(value, str) and "2026-27" in value for value in row)
    )
    header = rows[header_index]
    amount_col = _budget_column(header)
    label_col = 1
    heads: list[dict] = []
    for row in rows[header_index + 1:]:
        if len(row) <= amount_col or row[label_col] is None:
            continue
        label = _clean_label(str(row[label_col]))
        if not label or "total" in label.lower() or "opening balance" in label.lower():
            continue
        value = row[amount_col]
        if not isinstance(value, (int, float)) or value <= 0:
            continue
        heads.append({"label": label, "amount_cr": round(float(value) / 100, 4)})
    return sorted(heads, key=lambda row: row["amount_cr"], reverse=True)


def _clean_label(label: str) -> str:
    cleaned = label.replace("–", "-").replace("—", "-")
    cleaned = re.sub(r"^\s*\d+\s*[-.]?\s*", "", cleaned)
    cleaned = re.sub(r"^\s*&\s*", "", cleaned)
    cleaned = cleaned.replace(" & ", " and ")
    return " ".join(cleaned.split()).strip()
